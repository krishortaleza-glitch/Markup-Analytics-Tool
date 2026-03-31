import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Wholesale Markup Analytics", layout="wide")
st.title("💰 Wholesale Markup Analytics Tool")

@st.cache_data
def load_file(file):
    if file.name.endswith(".csv"):
        return pd.read_csv(file)
    return pd.read_excel(file)

st.header("Upload Files")

inv_file = st.file_uploader("Invoices", type=["xlsx", "csv"])
prod_file = st.file_uploader("Products File", type=["xlsx", "csv"])
front_file = st.file_uploader("Frontline", type=["xlsx", "csv"])
tax_file = st.file_uploader("Taxes", type=["xlsx", "csv"])
store_file = st.file_uploader("Storelist", type=["xlsx", "csv"])

if inv_file and prod_file and front_file and tax_file and store_file:

    inv = load_file(inv_file)
    prod = load_file(prod_file)
    front = load_file(front_file)
    tax = load_file(tax_file)
    store = load_file(store_file)

    st.success("Files loaded")

    col1, col2, col3, col4, col5 = st.columns(5)

    with col1:
        inv_store = st.selectbox("Invoice Store", inv.columns)
        inv_product = st.selectbox("Invoice ProductID", inv.columns)
        inv_cost = st.selectbox("Invoice Cost", inv.columns)

    with col2:
        prod_id = st.selectbox("Products ProductID", prod.columns)
        prod_family = st.selectbox("Products Family", prod.columns)

    with col3:
        front_family = st.selectbox("Frontline Family", front.columns)
        front_cost = st.selectbox("Frontline Cost", front.columns)
        front_start = st.selectbox("Start Date", front.columns)
        front_end = st.selectbox("End Date", front.columns)

    with col4:
        tax_state = st.selectbox("Tax State", tax.columns)
        tax_value = st.selectbox("Tax Value", tax.columns)

    with col5:
        store_store = st.selectbox("Storelist Store", store.columns)
        store_state = st.selectbox("Storelist State", store.columns)

    if st.button("🚀 Run Analysis"):

        progress = st.progress(0)
        status = st.empty()

        # ==============================
        # CLEAN DATA
        # ==============================
        inv["ProductID_clean"] = inv[inv_product].astype(str).str.strip()
        prod["ProductID_clean"] = prod[prod_id].astype(str).str.strip()

        prod[prod_family] = prod[prod_family].astype(str).str.upper().str.strip()
        front[front_family] = front[front_family].astype(str).str.upper().str.strip()

        inv["store_clean"] = inv[inv_store].astype(str).str.strip()
        store["store_clean"] = store[store_store].astype(str).str.strip()

        store["state_clean"] = store[store_state].astype(str).str.strip()
        tax["state_clean"] = tax[tax_state].astype(str).str.strip()

        progress.progress(10)

        # ==============================
        # MAP PRODUCT → FAMILY
        # ==============================
        status.text("Mapping Product → Family...")
        merged = inv.merge(
            prod[["ProductID_clean", prod_family]],
            on="ProductID_clean",
            how="left"
        )

        progress.progress(25)

        # ==============================
        # ACTIVE FRONTLINE
        # ==============================
        status.text("Filtering active frontline...")
        today = pd.Timestamp.today()

        front[front_start] = pd.to_datetime(front[front_start], errors="coerce")
        front[front_end] = pd.to_datetime(front[front_end], errors="coerce")
        front[front_end] = front[front_end].fillna(pd.Timestamp.max)

        active_front = front[
            (front[front_start] <= today) &
            (front[front_end] >= today)
        ]

        # Latest per family
        active_front = (
            active_front
            .sort_values(front_start, ascending=False)
            .drop_duplicates(subset=[front_family])
        )

        progress.progress(45)

        # ==============================
        # 🔥 FAST PARTIAL MATCH (VECTOR)
        # ==============================
        status.text("Vectorized family matching...")

        # Create mapping table
        families = merged[prod_family].dropna().unique()
        front_vals = active_front[[front_family, front_cost]]

        mapping = []

        for fam in families:
            match = front_vals[
                front_vals[front_family].str.contains(fam, na=False)
            ]
            if len(match) > 0:
                mapping.append((fam, match.iloc[0][front_cost]))

        map_df = pd.DataFrame(mapping, columns=[prod_family, "Frontline"])

        merged = merged.merge(map_df, on=prod_family, how="left")

        progress.progress(65)

        # ==============================
        # STORE → STATE
        # ==============================
        merged = merged.merge(
            store[["store_clean", "state_clean"]],
            on="store_clean",
            how="left"
        )

        progress.progress(75)

        # ==============================
        # TAX
        # ==============================
        merged = merged.merge(
            tax[["state_clean", tax_value]],
            on="state_clean",
            how="left"
        )

        progress.progress(85)

        # ==============================
        # CALCULATIONS
        # ==============================
        merged["State"] = merged["state_clean"]
        merged["Family"] = merged[prod_family]

        merged["Invoice Cost"] = pd.to_numeric(merged[inv_cost], errors="coerce")
        merged["Frontline"] = pd.to_numeric(merged["Frontline"], errors="coerce")
        merged["Tax"] = pd.to_numeric(merged[tax_value], errors="coerce")

        merged["Tax"] = merged["Tax"].apply(
            lambda x: x/100 if pd.notna(x) and x > 1 else x
        )

        merged["Frontline"] = merged["Frontline"].fillna(0)
        merged["Tax"] = merged["Tax"].fillna(0)

        merged["Total Cost"] = merged["Frontline"] * (1 + merged["Tax"])
        merged["Markup"] = merged["Invoice Cost"] - merged["Total Cost"]

        merged["Markup %"] = merged["Markup"] / merged["Total Cost"]
        merged["Markup %"] = merged["Markup %"].replace([float("inf"), -float("inf")], 0)

        progress.progress(90)

        # ==============================
        # FREQUENCY
        # ==============================
        freq = (
            merged
            .dropna(subset=["State", "Family", "Invoice Cost"])
            .groupby(["State", "Family", "Invoice Cost"])
            .size()
            .reset_index(name="Frequency")
        )

        freq["Top"] = (
            freq.groupby(["State", "Family"])["Frequency"]
            .transform("max") == freq["Frequency"]
        )

        merged = merged.merge(freq, on=["State", "Family", "Invoice Cost"], how="left")

        # ==============================
        # OUTPUT
        # ==============================
        final = merged[[
            "State","Family","Invoice Cost","Frontline","Tax",
            "Total Cost","Markup","Markup %","Frequency","Top"
        ]]

        output = BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            final.to_excel(writer, sheet_name="Analysis", index=False)
            merged.to_excel(writer, sheet_name="Full Output", index=False)

        output.seek(0)

        wb = load_workbook(output)
        ws = wb["Analysis"]

        green = PatternFill(start_color="C6EFCE", fill_type="solid")
        top_col = list(final.columns).index("Top") + 1

        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=top_col).value:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = green

        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        progress.progress(100)
        st.success("✅ Done!")

        st.download_button(
            "📥 Download Analysis",
            data=final_output,
            file_name=f"markup_analysis_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )
